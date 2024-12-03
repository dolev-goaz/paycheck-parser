import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
import os

DEFAULT_SHEET_NAME = "main_sheet"

def month_str_to_period(month: str):
    return pd.Period(f"20{month[-2:]}-{month[:2]}", freq="M")

def open_or_create_dataframe(excel_file_path: str, sheet_name: str):
    try:
        # Load the existing workbook and the specific sheet
        workbook = load_workbook(excel_file_path)
        if sheet_name in workbook.sheetnames:
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, index_col=0, engine="openpyxl")
        else:
            df = pd.DataFrame()  # If sheet doesn't exist, create a new DataFrame
    except:
        workbook = None
        df = pd.DataFrame()  # If file doesn't exist, create a new DataFrame

    return df, workbook

def column_fit_width(worksheet, column_id: str):
    component_column = worksheet[column_id]
    max_length = 0
    for cell in component_column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    worksheet.column_dimensions[column_id].width = max_length + 2 # padding

def centralize_cells(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            # Set horizontal and vertical alignment to center
            cell.alignment = Alignment(horizontal="center", vertical="center")

def mark_color_changes(excel_file_path: str, sheet_name = DEFAULT_SHEET_NAME):
    df, workbook = open_or_create_dataframe(excel_file_path, sheet_name)
    worksheet = workbook[sheet_name]
    for month_column_idx in range(1, len(df.columns)):
        # indices start from 1 dont ask me
        
        month_column_idx += 2 # ignore first two columns- 'component' column and first column with nothing to compare to
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green fill for increased values
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Red fill for decreased values
        green_font = Font(color="2C612E")
        red_font = Font(color="9C0055")
        
        # First month, nothing to compare to
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=0, max_col=month_column_idx):
            current_value, previous_value = row[month_column_idx-1].value, row[month_column_idx-2].value

            if pd.isna(previous_value):
                continue  # Skip if no previous value (for example, the first month)

            # Check if the value has increased or decreased
            if current_value > previous_value:
                row[month_column_idx-1].fill = green_fill  # Set green if the value increased
                row[month_column_idx-1].font = green_font
            elif current_value < previous_value:
                row[month_column_idx-1].fill = red_fill  # Set red if the value decreased
                row[month_column_idx-1].font = red_font
    workbook.save(excel_file_path)




def update_paycheck_log(excel_file_path: str, month: str, paycheck_data: dict[str, ], sheet_name=DEFAULT_SHEET_NAME):
    df, _ = open_or_create_dataframe(excel_file_path, sheet_name)
    
    df.columns = pd.to_datetime(df.columns).to_period("M")
    month_period = month_str_to_period(month)
    if month_period not in df.columns:
        df[month_period] = 0.0
    
    # Update the DataFrame with paycheck data for the given month
    for component, value in paycheck_data.items():
        if component not in df.index:
            df.loc[component] = [0.0] * len(df.columns)  # Add new row with NaN values
        df.loc[component, month_period] = pd.to_numeric(value, errors='coerce')
    
    # Write back to Excel
    write_mode = "a" if os.path.exists(excel_file_path) else "w"
    with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode=write_mode) as writer:
        if sheet_name in writer.book.sheetnames:
            del writer.book[sheet_name]
        df.to_excel(writer, sheet_name=sheet_name)
        # Align all cells in the sheet to the center
        worksheet = writer.sheets[sheet_name]  # Get the worksheet object
        
        centralize_cells(worksheet)
        column_fit_width(worksheet, 'A')
        
def insert_header(excel_file_path: str, header: str, row: int, sheet_name = DEFAULT_SHEET_NAME):
    _, workbook = open_or_create_dataframe(excel_file_path, sheet_name)
    worksheet = workbook[sheet_name]
    
    worksheet.insert_rows(row)
    worksheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=worksheet.max_column)
    header_cell = worksheet.cell(row=row, column=1)
    header_cell.value = header
    header_cell.font = Font(bold=True, size=12)  # Make the font bold
    header_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center the text
    header_cell.fill = PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")    # Red fill for decreased values
    worksheet.row_dimensions[row].height = 22
    workbook.save(excel_file_path)